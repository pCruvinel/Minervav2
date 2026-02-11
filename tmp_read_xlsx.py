import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

base = r"c:\Users\Usuario\OneDrive\Documentos\claude\Minervav2\docs\ordens-de-servico\WORKFLOWS\SPCI e SPDA"

for fname in ["MODELO gerado - inspeção de SPCI.xlsx", "MODELO gerado - inspeção de SPDA.xlsx"]:
    path = os.path.join(base, fname)
    print(f"\n{'='*80}")
    print(f"FILE: {fname}")
    print(f"{'='*80}")
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ---")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            vals = []
            for cell in row:
                v = cell.value
                if v is not None:
                    vals.append(f"[{cell.column_letter}{cell.row}] {v}")
            if vals:
                print(" | ".join(vals))
    wb.close()
